#!/usr/bin/python
# -*- coding: utf-8 -*-
from openpyxl import load_workbook

import logging
from optparse import OptionParser
from openerp import models, fields, api
from factory_csv import FactoryCSV
import time

logging.basicConfig()
_logger = logging.getLogger('SCRIPT CREAR PLANTILLAS DE PLANE INTERNO')

def main():
    usage = "create template file csv for plan interno: %prog [options]"
    parser = OptionParser(usage)

    parser.add_option("-f", "--fiel", dest="fiel", help="source file")
    parser.add_option("-d", "--debug", dest="debug", help="Mostrar mensajes de debug utilize 10", default=10)

    (options, args) = parser.parse_args()
    _logger.setLevel(int(options.debug))

    if not options.fiel:
        parser.error('Parametro fiel no especificado')

    # Inicio
    factory_csv = FactoryCSV()
    factory_csv.create_file_csv()

    # leer archivo
    documento = load_workbook("file/" + options.fiel)
    name_sheet = documento.get_sheet_names()

    _logger.debug('\n')
    _logger.debug('**************************************')
    _logger.debug('*** Hoja de Trbajo: ' + name_sheet[0] + ' ***')
    _logger.debug('**************************************')

    # obtener primera y unica hoja de trabajo en el archivo xls
    get_sheet = documento.get_sheet_by_name(name_sheet[0])
    row_cont=7
    hoy = (time.strftime("%d/%m/%y"))
    for i in range(get_sheet.max_row):
        row_cont += 1
        # Plan
        no = get_sheet.cell(row=row_cont, column=1)
        fecha = get_sheet.cell(row=row_cont, column=2)
        origen = get_sheet.cell(row=row_cont, column=3)
        proceso_origen = get_sheet.cell(row=row_cont, column=4)
        dependencia = get_sheet.cell(row=row_cont, column=5)
        responsable = get_sheet.cell(row=row_cont, column=18) # coordinar con camilo
        orfeo = get_sheet.cell(row=row_cont, column=19)
        responsable_oci_usuario = get_sheet.cell(row=row_cont, column=24)
        responsable_oci_gmail = get_sheet.cell(row=row_cont, column=25)
        responsable_oci_nombre = get_sheet.cell(row=row_cont, column=26)

        # Hallazgo
        hallazgo = get_sheet.cell(row=row_cont, column=6)
        causa_hallaz = get_sheet.cell(row=row_cont, column=7)

        # Acción
        accion = get_sheet.cell(row=row_cont, column=8)
        objetivo = get_sheet.cell(row=row_cont, column=9)
        indicador = get_sheet.cell(row=row_cont, column=10)
        meta = get_sheet.cell(row=row_cont, column=11)
        unidad_medida = get_sheet.cell(row=row_cont, column=12)
        area_resp = get_sheet.cell(row=row_cont, column=13)
        recursos = get_sheet.cell(row=row_cont, column=14)
        fecha_inicio = get_sheet.cell(row=row_cont, column=15)
        fecha_fin = get_sheet.cell(row=row_cont, column=16)
        ejecutor_usuario = get_sheet.cell(row=row_cont, column=27)
        ejecutor_nombre = get_sheet.cell(row=row_cont, column=28)
        ejecutor_gmail = get_sheet.cell(row=row_cont, column=29)

        # Avance
        avance = get_sheet.cell(row=row_cont, column=17)
        calificacion_cumplida = get_sheet.cell(row=row_cont, column=21)
        calificacion_no_cumplida = get_sheet.cell(row=row_cont, column=22)
        calificacion_en_tiempo = get_sheet.cell(row=row_cont, column=23)
        datos_avances = factory_csv.get_tipo_calificacion_avances_internos(calificacion_cumplida.value, calificacion_no_cumplida.value, calificacion_en_tiempo.value)

        # insertar primer plan
        if i == 0:
            # primer plan
            factory_csv.add_line_template_plan(
                origen.value, 'interno', orfeo.value,
                fecha.value, dependencia.value, responsable_oci_usuario.value, origen.value,
                'sub_origen', proceso_origen.value, responsable_oci_nombre.value,
                responsable_oci_gmail.value, 'OCI', '1'
            )
            # primer hallazgo
            factory_csv.add_line_template_hallazgo(factory_csv.get_id_plan(origen.value), responsable_oci_usuario.value, hallazgo.value[:15],
                dependencia.value, hallazgo.value, causa_hallaz.value
            )
            # primer Acción
            factory_csv.add_line_template_accion(factory_csv.get_id_hallazgo(hallazgo.value), 'preventivo', responsable_oci_usuario.value, area_resp.value,  # ?? cordinar tipo de accion con camilo
                accion.value, objetivo.value, indicador.value, unidad_medida.value, meta.value, recursos.value,
                fecha_inicio.value, fecha_fin.value, ejecutor_usuario.value, ejecutor_nombre.value, ejecutor_gmail.value, '2') # el 2 es para asiganr el grupo de ejcutor

        # buscamos existencia de plan
        if origen.value and hallazgo.value and accion.value:
            if not factory_csv.find_existing_plan(origen.value):
                # Crear Plan # ?? cordinar con camilo el sub-origen
                factory_csv.add_line_template_plan(
                    origen.value, 'interno', orfeo.value,
                    fecha.value, dependencia.value, responsable_oci_usuario.value, origen.value,
                    'sub_origen', proceso_origen.value, responsable_oci_nombre.value,
                    responsable_oci_gmail.value, 'OCI', '1'
                )
            if not factory_csv.find_existing_hallazgo(hallazgo.value):
                #Crear Hallazgo
                factory_csv.add_line_template_hallazgo(factory_csv.get_id_plan(origen.value), responsable_oci_usuario.value, hallazgo.value[:15],
                    dependencia.value, hallazgo.value, causa_hallaz.value
                )
            if not factory_csv.find_existing_accion(accion.value):
                #Crear Acción
                factory_csv.add_line_template_accion(factory_csv.get_id_hallazgo(hallazgo.value), 'preventivo', responsable_oci_usuario.value, area_resp.value,
                    accion.value, objetivo.value, indicador.value, unidad_medida.value, meta.value, recursos.value,
                    fecha_inicio.value, fecha_fin.value, ejecutor_usuario.value, ejecutor_nombre.value, ejecutor_gmail.value, '2'
                )
            factory_csv.add_line_template_avances(factory_csv.get_id_accion(accion.value), avance.value, hoy,
                datos_avances[0], datos_avances[2], datos_avances[1]
            )

    _logger.debug('\n')
    _logger.debug('**************************************')
    _logger.debug('*** Fin Hoja de Trbajo: ' + name_sheet[0] + ' ***')
    _logger.debug('**************************************')
    # fin for

if __name__ == '__main__':
    main()